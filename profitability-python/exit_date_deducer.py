import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime, date

# ============================================================
# END-TO-END SCRIPT (Next-bar Open entry + DEBUG)
# Input  : /mnt/data/RSID_GA_NIFTY_Daily_Observations.xlsx
# Output : /mnt/data/RSID_GA_NIFTY_Daily_Observations_exit_autofill_v2_profit_deltas_all_nextopen.xlsx
# ============================================================

INPUT_XLSX  = "RSID_GA_NIFTY_Daily_Observations.xlsx"
OUTPUT_XLSX = "RSID_GA_NIFTY_Daily_Observations_exit_autofill_v2_profit_deltas_all_nextopen.xlsx"

DEBUG = True

# IMPORTANT: this makes the script overwrite Exit_Date_Analyst even if you filled it manually
# (so the very first row will be recomputed too)
FORCE_RECALC_EXITS = True

# Dataset config
DATA_ROWS = 123              # number of benchmark rows
MAX_HOLD_BARS = 120

# Exit rule thresholds (v2)
SELL_RSI_PRIMARY = 35.0
SELL_RSI_RELAX   = 40.0
BUY_RSI_PRIMARY  = 60.0
BUY_RSI_RELAX    = 55.0
BUY_RSI_MINRELAX = 50.0

def to_date_safe(x):
    if x is None:
        return None
    if isinstance(x, (datetime, date)):
        return x.date() if isinstance(x, datetime) else x
    s = str(x).strip()
    if s == "" or s.upper() in ("N/A", "NA", "NONE"):
        return None
    try:
        return pd.to_datetime(x).date()
    except Exception:
        return None

def wilder_rsi(close: pd.Series, period: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = (-delta).clip(lower=0)
    avg_gain = gain.ewm(alpha=1/period, adjust=False, min_periods=period).mean()
    avg_loss = loss.ewm(alpha=1/period, adjust=False, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))

def ensure_column(ws, col_map, name):
    if name not in col_map:
        ws.cell(1, ws.max_column + 1).value = name
        col_map[name] = ws.max_column
    return col_map[name]

def set_ddmmyyyy_format(ws, col_indexes):
    for r in range(2, ws.max_row + 1):
        for c in col_indexes:
            cell = ws.cell(r, c)
            v = cell.value
            if isinstance(v, (datetime, date)):
                cell.number_format = "dd/mm/yyyy"

# ---------------------------
# Load workbook
# ---------------------------
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["NIFTY_Daily"]
ws_p = wb["Prices"]

dbg = {
    "original_rows_in_sheet": ws.max_row,
    "trimmed_to_rows": None,
    "exits_written": 0,
    "exit_maxhold": 0,
    "exit_rsi_rule": 0,
    "profits_manual_written": 0,
    "profits_sld_written": 0,
    "profits_tv_written": 0,
    "profits_tli_written": 0,
    "deltas_sld_vs_manual_written": 0,
    "deltas_sld_vs_tv_written": 0,
    "deltas_tli_vs_manual_written": 0,
    "rows_missing_required_fields": 0,
}

# ---------------------------
# Load Prices
# ---------------------------
price_rows = []
for row in ws_p.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    price_rows.append(row)

prices = pd.DataFrame(price_rows, columns=["Date", "Open", "High", "Low", "Close"])
prices["Date"] = pd.to_datetime(prices["Date"]).dt.date
prices = prices.sort_values("Date").reset_index(drop=True)

prices["RSI14"] = wilder_rsi(prices["Close"], 14)

date_list = prices["Date"].tolist()
date_to_idx = {d: i for i, d in enumerate(date_list)}
date_to_open = prices.set_index("Date")["Open"].to_dict()
date_to_close = prices.set_index("Date")["Close"].to_dict()

def next_bar_open(signal_date: date):
    """Open of the next trading day after signal_date."""
    if signal_date not in date_to_idx:
        return None
    i = date_to_idx[signal_date] + 1
    if i >= len(date_list):
        return None
    return date_to_open.get(date_list[i])

def profit_nextopen(signal_date: date, exit_date: date, is_sell: bool):
    """Entry: next-bar open after signal date; Exit: close on exit date."""
    if signal_date is None or exit_date is None:
        return None
    entry_open = next_bar_open(signal_date)
    exit_close = date_to_close.get(exit_date)
    if entry_open is None or exit_close is None:
        return None
    p = (entry_open - exit_close) / entry_open if is_sell else (exit_close - entry_open) / entry_open
    return round(p * 100, 3)

# ---------------------------
# Column map
# ---------------------------
headers = [c.value for c in ws[1]]
col = {h: i + 1 for i, h in enumerate(headers)}

H_CONFIRM     = "Confirmation Date"
H_TRADE       = "Trade Type"
H_EXIT_DATE   = "Exit_Date_Analyst"
H_EXIT_PRICE  = "Exit_Price_Analyst"
H_EXIT_REASON = "Exit_Reason"

H_SLD_DATE    = "SLD_Date"
H_TV_DATE     = "T1_TV_Date "        # note trailing space in your workbook
H_TLI_DATE    = "TLI_Date (TTLI)"

# Ensure required output columns exist
for h in [
    H_EXIT_DATE, H_EXIT_PRICE, H_EXIT_REASON,
    "Profit_%", "Profit_SLD_%", "Profit_TV_%", "Profit_TLI_%",
    "Delta_SLD_vs_Manual_%", "Delta_SLD_vs_TV_%", "Delta_TLI_vs_Manual_%"
]:
    ensure_column(ws, col, h)

# ---------------------------
# Step 1: Trim to 123 rows + header
# ---------------------------
target_max_row = 1 + DATA_ROWS
if ws.max_row > target_max_row:
    ws.delete_rows(target_max_row + 1, ws.max_row - target_max_row)
dbg["trimmed_to_rows"] = ws.max_row

# ---------------------------
# Step 2: Recompute Exit fields (overwrite all if FORCE_RECALC_EXITS)
# ---------------------------
for r in range(2, ws.max_row + 1):

    conf_val = ws.cell(r, col[H_CONFIRM]).value
    trade_val = ws.cell(r, col[H_TRADE]).value
    if conf_val is None or trade_val is None:
        dbg["rows_missing_required_fields"] += 1
        continue

    if not FORCE_RECALC_EXITS:
        existing_exit = ws.cell(r, col[H_EXIT_DATE]).value
        if existing_exit not in (None, "", "NA", "N/A"):
            continue

    conf_date = to_date_safe(conf_val)
    if conf_date is None or conf_date not in date_to_idx:
        dbg["rows_missing_required_fields"] += 1
        continue

    trade = str(trade_val).strip().lower()
    is_sell = trade.startswith("s")
    is_buy  = trade.startswith("b")
    if not (is_sell or is_buy):
        dbg["rows_missing_required_fields"] += 1
        continue

    start_idx = date_to_idx[conf_date]
    conf_close = float(prices.at[start_idx, "Close"])

    search_start = min(start_idx + 1, len(prices) - 1)
    search_end   = min(start_idx + MAX_HOLD_BARS, len(prices) - 1)

    exit_idx = None
    reason = None

    if is_sell:
        for i in range(search_start, search_end + 1):
            rsi = prices.at[i, "RSI14"]
            if pd.isna(rsi):
                continue
            if (rsi <= SELL_RSI_PRIMARY) and (prices.at[i, "Close"] < conf_close):
                exit_idx = i
                reason = f"RSI≤{SELL_RSI_PRIMARY:g} & Close<Conf"
                break
        if exit_idx is None:
            for i in range(search_start, search_end + 1):
                rsi = prices.at[i, "RSI14"]
                if pd.isna(rsi):
                    continue
                if (rsi <= SELL_RSI_RELAX) and (prices.at[i, "Close"] < conf_close):
                    exit_idx = i
                    reason = f"RSI≤{SELL_RSI_RELAX:g} & Close<Conf"
                    break

    else:
        for i in range(search_start, search_end + 1):
            rsi = prices.at[i, "RSI14"]
            if pd.isna(rsi):
                continue
            if (rsi >= BUY_RSI_PRIMARY) and (prices.at[i, "Close"] > conf_close):
                exit_idx = i
                reason = f"RSI≥{BUY_RSI_PRIMARY:g} & Close>Conf"
                break
        if exit_idx is None:
            for i in range(search_start, search_end + 1):
                rsi = prices.at[i, "RSI14"]
                if pd.isna(rsi):
                    continue
                if (rsi >= BUY_RSI_RELAX) and (prices.at[i, "Close"] > conf_close):
                    exit_idx = i
                    reason = f"RSI≥{BUY_RSI_RELAX:g} & Close>Conf"
                    break
        if exit_idx is None:
            for i in range(search_start, search_end + 1):
                rsi = prices.at[i, "RSI14"]
                if pd.isna(rsi):
                    continue
                if (rsi >= BUY_RSI_MINRELAX) and (prices.at[i, "Close"] > conf_close):
                    exit_idx = i
                    reason = f"RSI≥{BUY_RSI_MINRELAX:g} & Close>Conf"
                    break

    if exit_idx is None:
        exit_idx = search_end
        reason = f"MaxHold({MAX_HOLD_BARS})"
        dbg["exit_maxhold"] += 1
    else:
        dbg["exit_rsi_rule"] += 1

    exit_date = prices.at[exit_idx, "Date"]
    exit_close = float(prices.at[exit_idx, "Close"])

    ws.cell(r, col[H_EXIT_DATE]).value = exit_date
    ws.cell(r, col[H_EXIT_PRICE]).value = exit_close
    ws.cell(r, col[H_EXIT_REASON]).value = reason
    dbg["exits_written"] += 1

# ---------------------------
# Step 3: Profits using NEXT-BAR OPEN entry for ALL methods
# ---------------------------
for r in range(2, ws.max_row + 1):
    trade_val = ws.cell(r, col[H_TRADE]).value
    conf_val  = ws.cell(r, col[H_CONFIRM]).value
    exit_val  = ws.cell(r, col[H_EXIT_DATE]).value

    if trade_val is None or conf_val is None or exit_val is None:
        continue

    trade = str(trade_val).strip().lower()
    is_sell = trade.startswith("s")

    conf_date = to_date_safe(conf_val)
    exit_date = to_date_safe(exit_val)

    # Manual
    p = profit_nextopen(conf_date, exit_date, is_sell)
    if p is not None:
        ws.cell(r, col["Profit_%"]).value = p
        dbg["profits_manual_written"] += 1

    # SLD
    sld_d = to_date_safe(ws.cell(r, col[H_SLD_DATE]).value) if H_SLD_DATE in col else None
    p = profit_nextopen(sld_d, exit_date, is_sell) if sld_d else None
    if p is not None:
        ws.cell(r, col["Profit_SLD_%"]).value = p
        dbg["profits_sld_written"] += 1

    # TV
    tv_d = to_date_safe(ws.cell(r, col[H_TV_DATE]).value) if H_TV_DATE in col else None
    p = profit_nextopen(tv_d, exit_date, is_sell) if tv_d else None
    if p is not None:
        ws.cell(r, col["Profit_TV_%"]).value = p
        dbg["profits_tv_written"] += 1

    # TLI
    tli_d = to_date_safe(ws.cell(r, col[H_TLI_DATE]).value) if H_TLI_DATE in col else None
    p = profit_nextopen(tli_d, exit_date, is_sell) if tli_d else None
    if p is not None:
        ws.cell(r, col["Profit_TLI_%"]).value = p
        dbg["profits_tli_written"] += 1

# ---------------------------
# Step 4: Deltas
# ---------------------------
for r in range(2, ws.max_row + 1):
    manual = ws.cell(r, col["Profit_%"]).value
    sld    = ws.cell(r, col["Profit_SLD_%"]).value
    tv     = ws.cell(r, col["Profit_TV_%"]).value
    tli    = ws.cell(r, col["Profit_TLI_%"]).value

    if manual is not None and sld is not None:
        ws.cell(r, col["Delta_SLD_vs_Manual_%"]).value = round(float(sld) - float(manual), 3)
        dbg["deltas_sld_vs_manual_written"] += 1

    if sld is not None and tv is not None:
        ws.cell(r, col["Delta_SLD_vs_TV_%"]).value = round(float(sld) - float(tv), 3)
        dbg["deltas_sld_vs_tv_written"] += 1

    if manual is not None and tli is not None:
        ws.cell(r, col["Delta_TLI_vs_Manual_%"]).value = round(float(tli) - float(manual), 3)
        dbg["deltas_tli_vs_manual_written"] += 1

# ---------------------------
# Step 5: Date formatting
# ---------------------------
date_cols = []
for h in [H_CONFIRM, H_SLD_DATE, H_TV_DATE, H_TLI_DATE, H_EXIT_DATE]:
    if h in col:
        date_cols.append(col[h])
set_ddmmyyyy_format(ws, date_cols)

# Save
wb.save(OUTPUT_XLSX)

if DEBUG:
    print("\n========== DEBUG SUMMARY ==========")
    for k, v in dbg.items():
        print(f"{k:30s}: {v}")
    print("===================================")
    print("Saved:", OUTPUT_XLSX)
